import os
import PyPDF2
import re
from openpyxl import Workbook

# Crear un nuevo archivo de Excel y seleccionar la hoja de trabajo
workbook = Workbook()
sheet = workbook.active

# Escribir los encabezados de las columnas
sheet['A1'] = "Id Pedido Cliente"
sheet['B1'] = "Codigo de Producto"
sheet['C1'] = "Lote"
sheet['D1'] = "Cantidad"
sheet['E1'] = "Caducidad"

# Definir la expresión regular para buscar el texto después de la palabra "Cantidad"
regex = re.compile(r'de vencimiento Cantidad\s+([^\n]+)')
barcodes = re.compile(r'barras:\s+([^\s]+)' )
idPedido = re.compile(r'Original\s+([0-9]+)')

# Recorrer todos los archivos PDF en la carpeta
folder_path = 'ruta de la carpeta donde se encuentran los archivos pdf'
all_pages_info = []
for filename in os.listdir(folder_path):
    if filename.endswith('.pdf'):
        # Abrir el archivo PDF en modo lectura binaria
        with open(os.path.join(folder_path, filename), 'rb') as pdf_file:
            # Crear un objeto PDFFileReader
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            # Obtener el número de páginas del archivo PDF
            num_pages = len(pdf_reader.pages)

            # Crear una lista vacía para almacenar la información de cada página
            pages_info = []
            barcode_info = []

            # Recorrer todas las páginas del archivo PDF
            for page_number in range(num_pages):
                # Obtener la página actual del archivo PDF
                page = pdf_reader.pages[page_number]

                # Obtener el contenido de la página actual
                page_content = page.extract_text()

                # Buscar el texto después de la palabra "Cantidad" en el contenido de la página actual
                matches = regex.findall(page_content)
                barcodesMatch = barcodes.findall(page_content)
                idMatch = idPedido.findall(page_content)

                if matches:
                    # Si se encontró una coincidencia, obtener la línea siguiente
                    next_line_regex = re.compile(r'Cantidad\s+[^\n]+\n(.+)')
                    next_line_matches = next_line_regex.findall(page_content)

                    if next_line_matches:
                        next_lines = next_line_matches[0]
                    else:
                        next_lines = ""

                    # Agregar la información a la lista de información de cada página
                    for match in matches:
                        pages_info.append(match + " " + next_lines)
                if barcodesMatch:
                    for barcode in barcodesMatch:
                        barcode_info.append(barcode)

            #  Unir los elementos de barcode_info al principio de cada lista de pages_info y luego los unir en una sola lista
            pages_info = [x + " " + y for x, y in zip(barcode_info, pages_info)]

            # Escribir los resultados en las filas correspondientes
            for row, info in enumerate(pages_info, start=len(all_pages_info) + 2):
                data = info.split()
                sheet.cell(row=row, column=1, value=idMatch[0])
                sheet.cell(row=row, column=2, value=data[0])
                sheet.cell(row=row, column=3, value=data[1])
                sheet.cell(row=row, column=4, value=data[5])
                sheet.cell(row=row, column=5, value=data[4])

            all_pages_info.extend(pages_info) 

# Guardar el archivo de Excel
workbook.save('output.xlsx')

print("Se ejecuto correctamente")

